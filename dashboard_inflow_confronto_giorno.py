import pandas as pd
import datetime as dt
import numpy as np
import shutil
import time
from os import listdir
from os.path import isfile, join
from os.path import getmtime, getctime
from directories import *


# Oggi
shutil.copy(dashboard_inflow_source + r"\inflow_prodotti\inflow_prodotti_ins_man.xlsx",
            dashboard_inflow_oggi)

shutil.copy(dashboard_inflow_source + r"\inflow_prodotti\inflow_prodotti_nuove_vendite_2023.xlsx",
            dashboard_inflow_oggi)

#Ieri viene riempito con il programma "dashboard_inflow"

# PO - Aggiungere manualmente sia "inflow_prodotti_ins_man" sia "inflow_prodotti_nuove_vendite_2023_mese_corrente"
shutil.copy(dashboard_inflow_source + r"\inflow_prodotti\inflow_prodotti_ins_man.xlsx",
            dashboard_inflow_po)
shutil.copy(dashboard_inflow_source + r"\po_giornaliero_temp\inflow_prodotti_po_giornaliero.xlsx",
            dashboard_inflow_po)
shutil.copy(dashboard_inflow_source+f"/po_giornaliero/inflow_prodotti_nuove_vendite_2023_mese_corrente.xlsx",
            dashboard_inflow_po)



dashboard_inflow_ieri_files = [f for f in listdir(dashboard_inflow_ieri) if isfile(join(dashboard_inflow_ieri, f))]
dashboard_inflow_oggi_files = [f for f in listdir(dashboard_inflow_oggi) if isfile(join(dashboard_inflow_oggi, f))]
dashboard_inflow_po_files = [f for f in listdir(dashboard_inflow_po) if isfile(join(dashboard_inflow_po, f))]


dfs = []
for file in dashboard_inflow_ieri_files:
    data = pd.read_excel(dashboard_inflow_ieri+"\\"+file, na_values="#", 
                         usecols=["Sales Director","Agente dell'ORDINE","Canale di VENDITA","Cliente Merce","Numero Ordine","Origine Ordine","Data Ins. Ordine (monitoraggio)","Raccolto"])
    data["Data Ultimo Agg."] = dt.date.fromtimestamp(getmtime(dashboard_inflow_ieri+"\\"+file))
    data["Folder"] = dashboard_inflow_ieri.split("\\")[-1]
    dfs.append(data)

ieri = pd.concat(dfs, axis=0, ignore_index=True)
ieri = ieri.query('`Canale di VENDITA` != "E-Commerce"')

dfs = []
for file in dashboard_inflow_oggi_files:
    data = pd.read_excel(dashboard_inflow_oggi+"\\"+file, na_values="#",  
                         usecols=["Sales Director","Agente dell'ORDINE","Canale di VENDITA","Cliente Merce","Numero Ordine","Origine Ordine","Data Ins. Ordine (monitoraggio)","Raccolto"])
    data["Data Ultimo Agg."] = dt.date.fromtimestamp(getmtime(dashboard_inflow_oggi+"\\"+file))
    data["Folder"] = dashboard_inflow_oggi.split("\\")[-1]
    dfs.append(data)

oggi = pd.concat(dfs, axis=0, ignore_index=True)
oggi = oggi.query('`Canale di VENDITA` != "E-Commerce"')

dfs = []
for file in dashboard_inflow_po_files:
    data = pd.read_excel(dashboard_inflow_po+"\\"+file, na_values="#",  
                         usecols=["Sales Director","Agente dell'ORDINE","Canale di VENDITA","Cliente Merce","Numero Ordine","Origine Ordine","Data Ins. Ordine (monitoraggio)","Raccolto"])
    data["Data Ultimo Agg."] = dt.date.fromtimestamp(getmtime(dashboard_inflow_po+"\\"+file))
    data["Folder"] = dashboard_inflow_po.split("\\")[-1]
    dfs.append(data)

po = pd.concat(dfs, axis=0, ignore_index=True)
po = po.query('`Canale di VENDITA` != "E-Commerce"')

df = pd.concat([ieri, oggi, po], axis=0, ignore_index=True)


now = dt.datetime.now()
curr_month = now.month
current_year = now.year

df = df.loc[(df["Data Ins. Ordine (monitoraggio)"].dt.year == current_year) & (df["Data Ins. Ordine (monitoraggio)"].dt.month == curr_month),:]


df_giorno = df.groupby([df["Data Ins. Ordine (monitoraggio)"],"Folder"], dropna=False)["Raccolto"].sum().unstack(1).reset_index()
df_giorno.fillna(0, inplace=True)
df_giorno["Oggi - Ieri"] = df_giorno["Oggi"] - df_giorno["Ieri"]
df_giorno["Po - Oggi"] = df_giorno["Po"] - df_giorno["Oggi"]

sum_row = df_giorno.select_dtypes('number').sum()
sum_df = pd.DataFrame([sum_row], columns=df_giorno.columns)
df_giorno = pd.concat([df_giorno,sum_df], axis=0, ignore_index=True)
df_giorno.iloc[-1, 0] = "Totale"


df_ordine = df.groupby([df["Numero Ordine"],"Folder", "Origine Ordine", "Sales Director", df["Data Ins. Ordine (monitoraggio)"]], dropna=False)["Raccolto"].sum().unstack(1).reset_index()
df_ordine.fillna(0, inplace=True)
df_ordine["Oggi - Ieri"] = df_ordine["Oggi"] - df_ordine["Ieri"]
df_ordine["Po - Oggi"] = df_ordine["Po"] - df_ordine["Oggi"]

sum_row = df_ordine.select_dtypes('number').sum()
sum_df = pd.DataFrame([sum_row], columns=df_giorno.columns)
df_ordine = pd.concat([df_ordine,sum_df], axis=0, ignore_index=True)
df_ordine.iloc[-1, 0] = "Totale"

confronti_file = dashboard_inflow_confronto + f"\\confronti_{now.date()}_{now.hour}-{now.minute}.xlsx"

with pd.ExcelWriter(confronti_file, datetime_format='YYYY-MM-DD') as writer:
    df_giorno.to_excel(writer, sheet_name="giorno", index=False)
    df_ordine.to_excel(writer, sheet_name="ordine", index=False)

#############################################################################################################
# Format rows

import openpyxl

# Apri il file Excel con openpyxl
workbook = openpyxl.load_workbook(confronti_file)

# Ottieni tutti i nomi dei worksheet nel file
sheet_names = workbook.sheetnames

# Itera su tutti i worksheet
for worksheet in sheet_names:
    # Ottieni il worksheet corrente
    worksheet = workbook[worksheet]

    # Cerca la riga contenente "Totale" e applica lo stile "bold"
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value == 'Totale':
                for c in row:
                    c.font = openpyxl.styles.Font(bold=True)


    for col in worksheet.columns:
        for cell in col:
            if cell.row > 1 and col[0].value != 'Data Ins. Ordine (monitoraggio)':
                cell.number_format = '#,##0'


# Salva il file Excel modificato
workbook.save(confronti_file)

#############################################################################################################
# Autofit columns

import win32com.client as client
excel = client.gencache.EnsureDispatch('Excel.Application')
wb = excel.Workbooks.Open(confronti_file)

for sheet in wb.Sheets:
      
    ws = wb.Worksheets(sheet.Name)
    ws.Columns.AutoFit()

wb.Save()
excel.Application.Quit()